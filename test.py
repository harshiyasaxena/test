import os
import sys
import time
import tkinter as tk
from tkinter import ttk, messagebox
import win32com.client
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook

def on_button_click():
    dropdown1_value = dropdown1.get()
    dropdown2_value = dropdown2.get()
    dropdown3_value = dropdown3.get()

    if not dropdown1_value or not dropdown2_value or not dropdown3_value:
        messagebox.showwarning("Input Error", "Dropdowns cannot be blank")
        return

    if dropdown2_value == dropdown3_value:
        messagebox.showwarning("Input Error", "Status cannot be same")
        return

    Main_code(dropdown1_value, dropdown2_value, dropdown3_value)


def Main_code(dropdown1_value, dropdown2_value, dropdown3_value):

    folder_path = r"C:\Planner_Queue_Status_Update"
    file_path = os.path.join(folder_path, "Planner_Queue_Status_Update.xlsx")

    if not os.path.isdir(folder_path):
        messagebox.showerror("Error", f"Folder not found: {folder_path}")
        sys.exit()

    if not os.path.exists(file_path):
        messagebox.showerror("Error", f"File not found: {file_path}")
        sys.exit()

    # Close Excel if open
    excel = win32com.client.Dispatch("Excel.Application")
    for workbook in excel.Workbooks:
        if workbook.Name == "Planner_Queue_Status_Update.xlsx":
            workbook.Close(SaveChanges=False)
            break

    workbook = load_workbook(file_path)

    # Reset report sheet
    if "Planner Queue Report" in workbook.sheetnames:
        workbook.remove(workbook["Planner Queue Report"])
    sheet2 = workbook.create_sheet("Planner Queue Report")
    sheet1 = workbook["INPUTS"]

    Part_nos = [row.value for row in sheet1["A"] if row.value is not None]

    if not Part_nos:
        messagebox.showerror("Error", "No part numbers found in INPUTS column A")
        sys.exit()

    report_index = 1

    # PLAYWRIGHT STARTS
    with sync_playwright() as p:
        browser = p.firefox.launch(headless=False)
        page = browser.new_page()

        page.goto("https://hehe.pl")

        # Click Continue button
        page.click("input[value='Continue']")

        # Menu Selection
        try:
            page.wait_for_selector("#oCMenu_top2_0", timeout=30000)
            page.click("#oCMenu_top2_0")
        except:
            pass

        page.click("#oCMenu_sub112")

        # Dropdown values
        page.select_option("#model", label=dropdown1_value)
        page.select_option("#status", label=dropdown2_value)

        for Part_no in Part_nos:
            page.fill("input[name='partNumber']", str(Part_no))
            page.keyboard.press("Enter")
            time.sleep(1)

            index = 0
            while True:
                selector = f"select[name='plannerQList[{index}].status']"
                if not page.query_selector(selector):
                    if index == 0:
                        sheet2[f"A{report_index}"] = f"{Part_no} does not exist in Status -> {dropdown2_value}"
                        report_index += 1
                    break

                try:
                    page.select_option(selector, label=dropdown3_value)
                    page.click("#bt_Save")
                    sheet2[f"A{report_index}"] = f"{Part_no} -> Status changed to {dropdown3_value}"
                    report_index += 1
                except:
                    sheet2[f"A{report_index}"] = f"Unable to change status for {Part_no}"
                    report_index += 1

                index += 1

        browser.close()

    workbook.save(file_path)
    os.startfile(file_path)
    messagebox.showinfo("SUCCESS", "Automation completed successfully")


# GUI
root = tk.Tk()
root.title("Planner_Queue_Status_Update")

label_banner = tk.Label(root, text="Planner Queue Status Update",
                        font=("Helvetica", 20, "bold"),
                        bg="#4A90E2", fg="white",
                        pady=20, padx=40)
label_banner.grid(row=0, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

style = ttk.Style()
style.configure("TCombobox", font=("Helvetica", 20), padding=5)

label1 = tk.Label(root, text="Select Model:", font=("Helvetica", 15, "bold"))
label1.grid(row=1, column=0, padx=10, pady=10)

label2 = tk.Label(root, text="Change Status From:", font=("Helvetica", 15, "bold"))
label2.grid(row=2, column=0, padx=10, pady=10)

label3 = tk.Label(root, text="Change Status To:", font=("Helvetica", 15, "bold"))
label3.grid(row=3, column=0, padx=10, pady=10)

dropdown1_values = ["", "A", "B", "C"]
dropdown1 = ttk.Combobox(root, values=dropdown1_values)
dropdown1.grid(row=1, column=1, padx=10, pady=10)

dropdown2_values = ["", "New", "Open"]
dropdown2 = ttk.Combobox(root, values=dropdown2_values)
dropdown2.grid(row=2, column=1, padx=10, pady=10)

dropdown3_values_map = {"New": ["New", "In work"]}

dropdown3 = ttk.Combobox(root)
dropdown3.grid(row=3, column=1, padx=10, pady=10)

def update_dropdown3(event):
    selected_status = dropdown2.get()
    dropdown3["values"] = dropdown3_values_map.get(selected_status, [])
    dropdown3.set("")

dropdown2.bind("<<ComboboxSelected>>", update_dropdown3)

button = tk.Button(root, text="Submit", command=on_button_click, font=("Helvetica", 12, "bold"))
button.grid(row=4, column=0, columnspan=2, pady=10)

root.mainloop()
